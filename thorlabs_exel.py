#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Feb  5 21:47:56 2021

@author: bernhard

"""

#%% INPUT

# file location
target_dir = '/Users/bernhard/Documents/Administration/LKB/Orders/20220118_Thorlabs_Yoonseok/'
file_path = target_dir + 'shoppingCart.xls'

# initial dicount given by website (which needs to be added again)
discount_init = 0.00 # (0.00 for orders below 5k, 0.02 for orders above !)

# LKB discount
discount_fin = 0.09 # generally 0.09

# shipping costs
shipping_cost = 20.80

# translation mode
trans_flag = 0 # 0 = google translate API / 1 = DeepL API (500k char / month) 

#%%

import sys
import os
import numpy as np
from datetime import date
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
  
# change working directory to file directory
os.chdir(os.path.dirname(__file__))
from nacres_from_thorlabs import nacres_from_thorlabs

# load translate API
if trans_flag:
    
    import deepl
    
    # read authentification key
    try:
        auth_file = open("auth_deepl.txt", "r")
        auth_key = auth_file.read()
    except FileNotFoundError:
        sys.exit("DeepL authentification key not found!!\nset trans_flag = 0 or provide key in separate auth_deepl.txt file")
        
    # initialize translator    
    translator_deepl = deepl.Translator(auth_key)
    
else:
    
    # import local modified google translate package
    from google_trans_new_local.google_trans_new import google_translator
    
    # initialize translator  
    translator_google = google_translator() 

# define cell boarders join appropriate cells
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def num2str(num):
    output = f'{num:.2f}'.replace('.',',')
    if len(output) > 6:
        pos = len(output) - 6
        output = output[:pos] + '.' + output[pos:]
    return output

# load input data - check first the existence of .xls (default from thorlabs site), 
# then the .xlsx (from original implementation of this code)
if file_path.split('.')[-1] == 'xls':
    # load with pandas if old .xls format
    import pandas as pd
    pd_wb = pd.read_excel(file_path)
    input_data = pd_wb.values.tolist()
elif file_path.split('.')[-1] == 'xlsx':
    # load with openpyxl if .xlsx
    input_wb = openpyxl.load_workbook(file_path)
    input_ws = input_wb.active
    input_data = list( input_ws.values )[1:]

# handle som weird stuffs with xlsx...
for idx in range(len(input_data)):
    if input_data[idx][0] is None:
        idx = idx-1
        break
nr_items = idx+1
print(f'Processing {nr_items} items from {file_path}')

# open output file
output_wb = openpyxl.load_workbook('thorlabsBC_template.xlsx')
output_ws = output_wb.active

# set date of today
today = date.today()
output_ws.cell(row=3, column=10).value = today.strftime("%d/%m/%Y")

# insert necessary rows in output file
first_row = 10;
if nr_items > 1:
    output_ws.insert_rows(first_row, amount=(nr_items-1))  

# loop through items
price_unit = np.zeros((nr_items,))
price_full_disc = np.zeros((nr_items,))
price_unit_disc = np.zeros((nr_items,))

for idx in range(nr_items):

    # get values from input list
    desc = input_data[idx][1]
    quantity = input_data[idx][3]
    price_unit[idx] = round(input_data[idx][4] / (1-discount_init), 2)
    
    # translate decription into french
    if trans_flag:
        trans_result = translator_deepl.translate_text(desc, target_lang="FR")
        desc_fr = trans_result.text
    else:
        desc_fr = translator_google.translate(desc, lang_src='en', lang_tgt='fr')
        
    # calulate untaxed, dicount and full price
    price_unit_disc[idx] = round((1-discount_fin) * price_unit[idx], 2)
    price_full_disc[idx] = round(quantity * price_unit_disc[idx], 2)
    
    # if available from the dict, fill the NACRES code from the thorlabs code
    # input_data[idx][0] is the key, which is processsed through the nacres_from_thorlabs()
    nacres = nacres_from_thorlabs(input_data[idx][0], input_data[idx][1])
    
    # fill output file entries
    output_ws.cell(row=first_row+idx, column=1).value = quantity
    output_ws.cell(row=first_row+idx, column=2).value = desc_fr
    output_ws.cell(row=first_row+idx, column=5).value = input_data[idx][0]
    output_ws.cell(row=first_row+idx, column=7).value = nacres
    output_ws.cell(row=first_row+idx, column=8).value = num2str(price_unit[idx])
    output_ws.cell(row=first_row+idx, column=9).value = num2str(discount_fin*100)
    output_ws.cell(row=first_row+idx, column=10).value = num2str(price_full_disc[idx])
    
    # fix cell alignment
    output_ws.cell(row=first_row+idx, column=8).alignment = Alignment(horizontal='right')
    output_ws.cell(row=first_row+idx, column=9).alignment = Alignment(horizontal='right')
    output_ws.cell(row=first_row+idx, column=10).alignment = Alignment(horizontal='right')
    
    # fix cell boarders join appropriate cells
    for idx_col in range(1,11):
        output_ws.cell(row=first_row+idx, column=idx_col).border = thin_border
    output_ws.merge_cells(start_row=first_row+idx, start_column=2, end_row=first_row+idx, end_column=4)
    output_ws.merge_cells(start_row=first_row+idx, start_column=5, end_row=first_row+idx, end_column=6)

# insert the shipping cost
output_ws.cell(row=first_row+nr_items, column=10).value = num2str(shipping_cost)
output_ws.cell(row=first_row+nr_items, column=10).alignment = Alignment(horizontal='right')

# calculate total price    
price_nr_brut_disc_sum = round(np.sum(price_full_disc) + shipping_cost, 2)

# insert total amount in output file
output_ws.cell(row=first_row+nr_items+1, column=10).value = num2str(price_nr_brut_disc_sum)
output_ws.cell(row=first_row+nr_items+1, column=10).alignment = Alignment(horizontal='right')
    
# save output file
output_file = target_dir + 'thorlabsBC.xlsx'
output_wb.save(output_file)
print(f'Output written in {output_file}')
